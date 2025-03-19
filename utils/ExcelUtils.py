import openpyxl


def get_row_count(path,sheet_name):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name]
    return sheet.max_row


def get_column_count(path,sheet_name):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name]
    return sheet.max_column

def get_cell_data(path,sheet_name,row_number,column_number):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name]
    return sheet.cell(row=row_number,column=column_number).value

#
# def set_cel_data(path,sheet_name,row_number,columnnumber):
#     workbook=openpyxl.load_workbook(path)
#     sheet=workbook[sheet_name]
#     sheet.cell(row=row_number,column=columnnumber).value=cell_data
#     workbook.save

def get_data_from_excel(path,sheetname):
    final_list=[]
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]
    total_row=sheet.max_row
    total_column=sheet.max_column

    for r in range(2,total_row+1):
        row_list=[]
        for c in range(1,total_column+1):
            row_list.append(sheet.cell(row=r,column=c).value)
        final_list.append(row_list)
    return final_list
